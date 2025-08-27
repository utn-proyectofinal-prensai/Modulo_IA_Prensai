import pandas as pd
import requests
from bs4 import BeautifulSoup
import chardet
import os
import logging
import re
import unicodedata
import time
from datetime import datetime

# Levantar un logger
def setup_logger(filename):
    # Configura el logger para registrar eventos en un archivo.
    # Crear carpeta de logs si no existe
    os.makedirs('Logs', exist_ok=True)
    
    # Configurar el logger
    logging.basicConfig(
        filename=f'Logs/{filename}',  # <--- el archivo va a Logs/
        filemode='a',  # append, para no pisar logs anteriores
        format='%(asctime)s %(levelname)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',  # <-- esto elimina los milisegundos
        level=logging.INFO  # cambiado a INFO para ver logs de valoración
    )

#Export a excel
def exportar_df_a_excel(df, export_path):
    """
    Exporta el DataFrame a un archivo Excel en la ruta export_path.
    Crea el directorio si no existe y loguea errores.
    Aplica formato con cuadrícula (bordes) a todas las celdas.
    """
    # Extraer el directorio del path completo
    dir_path = os.path.dirname(export_path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)

    try:
        # Crear un writer de Excel para poder aplicar formato
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Datos')
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Importar estilos
            from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
            
            # Definir estilos
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Estilo para el encabezado
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formato al encabezado (fila 1)
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Aplicar bordes y formato a todas las celdas de datos
            for row in range(1, len(df) + 2):  # +2 para incluir header
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Aplicar alineación centrada a todas las celdas
                    if row > 1:  # Solo para datos, no para el header
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas automáticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logging.info(f"DataFrame exportado exitosamente a {export_path} con formato de cuadrícula")
    except Exception as e:
        logging.error(f"Error al exportar DataFrame a Excel ({export_path}): {e}") 

# Función para obtener el texto plano de un link, manejando encoding
def get_texto_plano_from_link(link):
    """
    Descarga el HTML del link y retorna un string con '[TÍTULO]: ... [BODY]: ...'.
    Si no encuentra los tags, retorna el texto plano general. Loguea errores.
    """
    try:
        r = requests.get(link, timeout=10)
        if r.status_code == 200:
            enc = r.encoding if r.encoding else 'utf-8'
            try:
                texto_bytes = r.content
                detected_enc = chardet.detect(texto_bytes)['encoding']
                enc = detected_enc if detected_enc else enc
                html = texto_bytes.decode(enc, errors='replace')
            except Exception as e:
                logging.warning(f"⚠️ Problema al decodificar HTML de {link}: {e}")
                html = r.text  # Fallback
            soup = BeautifulSoup(html, 'html.parser')

            # Título
            titulo = None
            span_titulo = soup.find("span", class_="titulo")
            if span_titulo and span_titulo.get_text(strip=True):
                titulo = span_titulo.get_text(strip=True)
            # Si no hay título específico, buscá por <title> de la página
            if not titulo:
                if soup.title:
                    titulo = soup.title.get_text(strip=True)

            # Body principal
            span_detalle = soup.find("span", class_="detalleFull")
            if span_detalle and span_detalle.get_text(strip=True):
                body = span_detalle.get_text(separator=' ', strip=True)
            else:
                body = soup.get_text(separator=' ', strip=True)

            # Construir el texto final para IA
            if titulo:
                return f"[TÍTULO]: {titulo}\n[BODY]: {body}"
            else:
                return body
        else:
            logging.warning(f"⚠️ Status code {r.status_code} al acceder a {link}")
            return None
    except requests.exceptions.ConnectionError as e:
        if "Connection refused" in str(e):
            logging.error(f"🚫 Error PERMANENTE (servidor caído): {link} - {e}")
        elif "Max retries exceeded" in str(e):
            logging.error(f"🔄 Error de reintentos agotados: {link} - {e}")
        else:
            logging.error(f"🌐 Error de conexión: {link} - {e}")
        return None
    except requests.exceptions.Timeout as e:
        logging.error(f"⏰ Timeout al acceder a {link}: {e}")
        return None
    except Exception as e:
        logging.error(f"❌ Excepción al descargar/parsing {link}: {e}")
        return None

#Funcion para obtener los LINKs de un archivo Excel que va importar el usuario en la PRIMER COLUMNA, PRIMER HOJA. 
def obtener_links_del_usuario_desde_excel(EXCEL_URL_PATH):
    """
    Lee el archivo Excel, obtiene la primera columna de la primer hoja (links).
    Devuelve solo los links que contengan 'ejes.com' (caso-insensitive).
    Loguea cuáles son descartados.
    """
    try:
        df_excel = pd.read_excel(EXCEL_URL_PATH, sheet_name=0)
        if df_excel.shape[1] == 0:
            logging.warning(f"El archivo {EXCEL_URL_PATH} no tiene columnas.")
            return None
        links = df_excel.iloc[:, 0]  # Primera columna
        if links.isnull().all():
            logging.warning(f"La primera columna del archivo {EXCEL_URL_PATH} está vacía o es nula.")
            return None

        # Filtrar solo los links de Ejes
        mask_ejes = links.str.contains('ejes.com', case=False, na=False)
        descartados = links[~mask_ejes]
        links_filtrados = links[mask_ejes]

        # Loguear los descartados
        for link in descartados:
            logging.info(f"Link descartado por no pertenecer a ejes.com: {link}")

        return links_filtrados.reset_index(drop=True)  # Opcional: resetear índice
    except Exception as e:
        logging.error(f"Error al leer Excel {EXCEL_URL_PATH}: {e}")
        return None

def procesar_link_robusto(link, tipo='texto', max_reintentos=3):
    """
    Procesa un link de forma robusta con reintentos automáticos
    
    Args:
        link: URL a procesar
        tipo: 'texto' o 'html'
        max_reintentos: número máximo de intentos (default: 3)
    
    Returns:
        - Si tipo='texto': texto plano extraído
        - Si tipo='html': objeto BeautifulSoup parseado
        - None si falla definitivamente después de todos los intentos
    """
    logging.info(f"🔄 Iniciando procesamiento robusto de {link} (tipo: {tipo}, max_reintentos: {max_reintentos})")
    
    for intento in range(max_reintentos):
        try:
            logging.info(f"🌐 Intento {intento + 1}/{max_reintentos} para {link}")
            
            if tipo == 'texto':
                resultado = get_texto_plano_from_link(link)
                if resultado:
                    logging.info(f"✅ Éxito en intento {intento + 1} para {link}")
                    return resultado
                else:
                    logging.warning(f"⚠️ Intento {intento + 1} falló (sin resultado) para {link}")
                    
            elif tipo == 'html':
                resultado = get_html_object_from_link(link)
                if resultado:
                    logging.info(f"✅ Éxito en intento {intento + 1} para {link}")
                    return resultado
                else:
                    logging.warning(f"⚠️ Intento {intento + 1} falló (sin resultado) para {link}")
            else:
                logging.error(f"❌ Tipo '{tipo}' no válido. Debe ser 'texto' o 'html'")
                return None
                
        except requests.exceptions.ConnectionError as e:
            if intento < max_reintentos - 1:
                delay = (2 ** intento) * 2  # 2, 4, 8 segundos
                logging.warning(f"🌐 Error de conexión en intento {intento + 1} para {link}. Reintentando en {delay}s... Error: {e}")
                time.sleep(delay)
            else:
                logging.error(f"❌ {link} falló definitivamente después de {max_reintentos} intentos por error de conexión. Error: {e}")
                return None
                
        except requests.exceptions.Timeout as e:
            if intento < max_reintentos - 1:
                delay = (2 ** intento) * 2
                logging.warning(f"⏰ Timeout en intento {intento + 1} para {link}. Reintentando en {delay}s... Error: {e}")
                time.sleep(delay)
            else:
                logging.error(f"❌ {link} falló definitivamente después de {max_reintentos} intentos por timeout. Error: {e}")
                return None
                
        except Exception as e:
            if intento < max_reintentos - 1:
                delay = (2 ** intento) * 2
                logging.warning(f"⚠️ Error general en intento {intento + 1} para {link}. Reintentando en {delay}s... Error: {e}")
                time.sleep(delay)
            else:
                logging.error(f"❌ {link} falló definitivamente después de {max_reintentos} intentos por error general. Error: {e}")
                return None
    
    logging.error(f"❌ {link} falló definitivamente después de {max_reintentos} intentos")
    return None

#Función para obtener el HTML de un link y devolerlo como un "objetito" para luego poder procesarlo y rellenar los campos de mi DF. 
def get_html_object_from_link(link):
    """
    Descarga el HTML desde el link, lo parsea y devuelve un objeto BeautifulSoup.
    Si falla, retorna None y loguea el error.
    """
    try:
        response = requests.get(link, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        return soup
    except requests.exceptions.ConnectionError as e:
        if "Connection refused" in str(e):
            logging.error(f"🚫 Error PERMANENTE (servidor caído): {link} - {e}")
        elif "Max retries exceeded" in str(e):
            logging.error(f"🔄 Error de reintentos agotados: {link} - {e}")
        else:
            logging.error(f"🌐 Error de conexión: {link} - {e}")
        return None
    except requests.exceptions.Timeout as e:
        logging.error(f"⏰ Timeout al obtener HTML de {link}: {e}")
        return None
    except requests.exceptions.HTTPError as e:
        logging.error(f"📡 Error HTTP {e.response.status_code} al obtener HTML de {link}: {e}")
        return None
    except Exception as e:
        logging.error(f"❌ Error general al parsear HTML de {link}: {e}")
        return None

##// JUEGO DE FUNCIONES para trabajar con el HTML_OBJ
#Título
def get_titulo_from_html_obj(html_obj):
    """
    Extrae el título principal desde el objeto BeautifulSoup.
    Busca primero el <span class="titulo">, y si no existe, busca <title>.
    Loguea advertencias si no encuentra el título o el HTML es inválido.
    Retorna None si no hay título.
    """
    try:
        if html_obj is None:
            logging.warning("Objeto HTML no válido al intentar extraer el título.")
            return None

        # Primero busca el título de la nota (más relevante para prensa)
        tag = html_obj.find("span", class_="titulo")
        if tag:
            return tag.get_text(strip=True)

        # Si no lo encuentra, busca el <title> de la página
        title_tag = html_obj.find('title')
        if title_tag and title_tag.text.strip():
            logging.warning("No se encontró <span class='titulo'>, usando <title> de la página.")
            return title_tag.text.strip()

        logging.warning("No se encontró ningún título en el HTML.")
        return None

    except Exception as e:
        logging.error(f"Error al extraer título del HTML: {e}")
        return None

#Fecha
def get_fecha_from_html_obj(soup):
    """
    Extrae la fecha en formato DD/MM/YYYY del <span class='canal'>.
    Devuelve la fecha en formato 'YYYY-MM-DD', o None si no la encuentra.
    Loguea advertencias y errores.
    """
    try:
        if soup is None:
            logging.warning("Objeto HTML no válido al intentar extraer la fecha.")
            return None

        canal = soup.find("span", class_="canal")
        if canal:
            texto = canal.get_text(strip=True)
            # Buscar patrón de fecha: 2 dígitos / 2 dígitos / 4 dígitos
            match = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
            if match:
                fecha_str = match.group(1)  # '04/04/2025'
                # Convertir a formato ISO
                d, m, y = fecha_str.split("/")
                return f"{y}-{m}-{d}"
            else:
                logging.warning("No se encontró fecha en <span class='canal'>.")
                return None
        else:
            logging.warning("No se encontró <span class='canal'> en el HTML.")
            return None
    except Exception as e:
        logging.error(f"Error al extraer fecha del HTML: {e}")
        return None

#Medio
def get_medio_from_html_obj(soup):
    """
    Extrae el medio del <span class='canal'>.
    Devuelve el nombre del medio o None.
    Mejorado: ahora captura cualquier texto entre la fecha y el primer guion.
    """
    try:
        if soup is None:
            logging.warning("Objeto HTML no válido al intentar extraer el medio.")
            return None

        canal = soup.find("span", class_="canal")
        if canal:
            texto = canal.get_text(strip=True)
            # Más flexible: captura lo que está entre la fecha y el primer guion
            match = re.search(r"\d{2}/\d{2}/\d{4}\s+([^-]+?)\s*-\s*", texto)
            if match:
                return match.group(1).strip()
            else:
                logging.warning(f"No se encontró el medio en <span class='canal'>. Texto: '{texto}'")
                return None
        else:
            logging.warning("No se encontró <span class='canal'> en el HTML.")
            return None
    except Exception as e:
        logging.error(f"Error al extraer medio del HTML: {e}")
        return None

#Soporte 
def get_soporte_from_html_obj(soup):
    """
    Determina el soporte a partir del medio extraído del HTML.
    Si contiene '.com' es 'WEB', si no, es 'GRÁFICA'.
    """
    try:
        medio = get_medio_from_html_obj(soup)
        if medio is None:
            logging.warning("No se pudo determinar el soporte porque no se encontró el medio.")
            return None
        if ".com" in medio.lower():
            return "WEB"
        else:
            return "GRÁFICA"
    except Exception as e:
        logging.error(f"Error al determinar soporte del HTML: {e}")
        return None

#Sección
def get_seccion_from_html_obj(soup):
    """
    Extrae la sección del <span class='canal'> si está disponible,
    o, si no, de la URL del primer <a href> relevante.
    Si no encuentra, devuelve 'Sitio' por defecto.
    """
    try:
        if soup is None:
            logging.warning("Objeto HTML no válido al intentar extraer la sección.")
            return "Sitio"

        # 1. Intentá método clásico (span canal)
        canal = soup.find("span", class_="canal")
        if canal:
            texto = canal.get_text(" ", strip=True)
            # ¡IMPORTANTE! Solo una barra:
            match = re.search(r"Nota\s*-\s*([^-]+?)(?:\s*-\s*Pag|\s*$)", texto)
            if match:
                return match.group(1).strip()

        # 2. Si no hay sección en <span class='canal'>, buscá en el primer <a href>
        a_tags = soup.find_all("a", href=True)
        for a in a_tags:
            url = a["href"]
            if any(x in url for x in ["infobae.com", "lanacion.com", "pagina12.com"]):
                m = re.search(r"\.com(?:\.ar)?/([^/]+)/", url)
                if m:
                    seccion = m.group(1)
                    seccion = seccion.replace("-", " ").capitalize()
                    return seccion
        # Si no se encontró nada, devuelve valor por defecto
        logging.warning("No se encontró sección en <span class='canal'> ni en los links <a href>. Asignando 'Sitio'.")
        return "Sitio"
    except Exception as e:
        logging.error(f"Error al extraer sección del HTML: {e}")
        return "Sitio"

#Cotización
def get_cotizacion_from_html_obj(soup):
    """
    Extrae la cotización de la nota desde los <span class='medicion'>.
    Devuelve el valor como string (ej: '$97.500') o None.
    """
    try:
        if soup is None:
            logging.warning("Objeto HTML no válido al intentar extraer cotización.")
            return None

        for span in soup.find_all("span", class_="medicion"):
            texto = span.get_text(strip=True)
            # Buscar línea con 'Cotización' o 'Cotizaci' (acentos pueden variar)
            if "Cotizaci" in texto:
                # Buscar monto en formato $XXX.XXX
                match = re.search(r"\$\s*([\d.,]+)", texto)
                if match:
                    return match.group(0)  # Incluye el signo $
        logging.warning("No se encontró cotización en <span class='medicion'>.")
        return None
    except Exception as e:
        logging.error(f"Error al extraer cotización del HTML: {e}")
        return None

#Alcance
def get_alcance_from_html_obj(soup):
    """
    Extrae el alcance del <span class='medicion'>.
    Devuelve el alcance o None.
    """
    try:
        if soup is None:
            logging.warning("Objeto HTML no válido al intentar extraer el alcance.")
            return None

        for span in soup.find_all("span", class_="medicion"):
            texto = span.get_text(strip=True)
            # Buscar línea con 'Audiencia'
            if "Audiencia:" in texto:
                # Buscar el valor después de "Audiencia:"
                match = re.search(r"Audiencia:\s*([^\s=]+)", texto)
                if match:
                    return match.group(1).strip()
        
        logging.warning("No se encontró alcance en <span class='medicion'>.")
        return None
    except Exception as e:
        logging.error(f"Error al extraer alcance del HTML: {e}")
        return None

def limpiar_autor(autor):
    """
    Limpia el autor removiendo prefijos comunes como "Por", "Por:", etc.
    
    Args:
        autor (str): Nombre del autor a limpiar
    
    Returns:
        str: Autor limpio
    """
    if not autor or pd.isnull(autor):
        return autor
    
    autor_limpio = autor.strip()
    
    # Lista de prefijos a remover
    prefijos = [
        "Por ",
        "Por: ",
        "POR ",
        "POR: ",
        "por ",
        "por: "
    ]
    
    # Remover prefijos
    for prefijo in prefijos:
        if autor_limpio.startswith(prefijo):
            autor_limpio = autor_limpio[len(prefijo):].strip()
            break
    
    return autor_limpio

def get_autor_from_html_obj(html_obj):
    """
    Extrae el autor desde el objeto BeautifulSoup.
    Busca el <span class="entrevistado"> que contiene el nombre del autor.
    SOLO es válido si aparece ANTES del <span class="detalleFull">.
    Si el span está vacío o no se encuentra, asigna "Redacción" como valor por defecto.
    Limpia prefijos comunes como "Por", "Por:", etc.
    Loguea advertencias si no encuentra el autor o el HTML es inválido.
    NUNCA devuelve None - siempre devuelve un valor.
    """
    try:
        if html_obj is None:
            logging.warning("Objeto HTML no válido al intentar extraer el autor. Asignando 'Redacción' por defecto.")
            return "Redacción"

        # Buscar el span detalleFull para validar posición
        detalle_full = html_obj.find("span", class_="detalleFull")
        
        # Buscar el autor en el span class="entrevistado"
        tag_entrevistado = html_obj.find("span", class_="entrevistado")
        
        if tag_entrevistado:
            # Verificar que aparezca antes del detalleFull
            if detalle_full and tag_entrevistado.find_previous_sibling("span", class_="detalleFull"):
                logging.warning("Span 'entrevistado' aparece después de 'detalleFull'. No es válido.")
                return "Redacción"
            
            autor = tag_entrevistado.get_text(strip=True)
            if autor:  # Si hay contenido, limpiarlo y devolverlo
                autor_limpio = limpiar_autor(autor)
                return autor_limpio
            else:
                # Si el span está vacío, asignar "Redacción"
                logging.info("Span 'entrevistado' vacío, asignando 'Redacción' por defecto.")
                return "Redacción"

        # Si no encuentra el span, asignar "Redacción"
        logging.info("No se encontró <span class='entrevistado'> válido en el HTML. Asignando 'Redacción' por defecto.")
        return "Redacción"

    except Exception as e:
        logging.error(f"Error al extraer autor del HTML: {e}. Asignando 'Redacción' por defecto.")
        return "Redacción"

##JUEGO DE FUNCIONES para trabajar con el HTML_OBJ //##
def normalizar_texto(texto):
    """
    Normaliza el texto removiendo acentos y convirtiendo a minúsculas.
    """
    try:
        if not texto or pd.isnull(texto):
            return ""
        
        # Normalizar unicode (NFD) y remover diacríticos
        texto_normalizado = unicodedata.normalize('NFD', texto.lower())
        # Remover caracteres diacríticos (acentos, diéresis, etc.)
        texto_sin_acentos = ''.join(c for c in texto_normalizado if not unicodedata.combining(c))
        return texto_sin_acentos
        
    except Exception as e:
        logging.error(f"Error al normalizar texto: {e}")
        return texto.lower() if texto else ""

def detectar_mencion(texto, palabra_clave):
    """
    Detecta si una palabra clave específica está mencionada en el texto.
    
    Args:
        texto (str): Texto plano a analizar
        palabra_clave (str): Palabra/entidad específica a buscar
    
    Returns:
        str: La palabra clave si se encuentra, string vacío si no
    """
    try:
        if not texto or pd.isnull(texto) or not palabra_clave or pd.isnull(palabra_clave):
            return ""
        
        texto_normalizado = normalizar_texto(texto)
        palabra_normalizada = normalizar_texto(palabra_clave.strip())
        
        # Buscar la palabra clave en el texto normalizado
        if palabra_normalizada in texto_normalizado:
            logging.debug(f"Mención detectada: {palabra_clave}")
            return palabra_clave.strip()
        else:
            return ""
        
    except Exception as e:
        logging.error(f"Error al detectar mención '{palabra_clave}' en el texto: {e}")
        return ""

def buscar_menciones(df, lista_menciones, max_menciones=5):
    """
    Busca menciones en el DataFrame y asigna los resultados a un solo campo 'MENCIONES' como lista.
    
    Args:
        df (DataFrame): DataFrame con la columna 'TEXTO_PLANO'
        lista_menciones (list): Lista de palabras/entidades a buscar
        max_menciones (int): Número máximo de menciones a procesar (por compatibilidad, no se usa)
    
    Returns:
        DataFrame: DataFrame con la columna 'MENCIONES' agregada
    """
    try:
        def encontrar_menciones_en_texto(texto):
            """
            Encuentra todas las menciones de la lista en un texto específico
            """
            if not texto or pd.isnull(texto):
                return []
            
            menciones_encontradas = []
            for palabra_clave in lista_menciones:
                if palabra_clave and not pd.isnull(palabra_clave):
                    resultado = detectar_mencion(texto, palabra_clave)
                    if resultado:  # Si se encontró la mención
                        menciones_encontradas.append(palabra_clave.strip())
            
            return menciones_encontradas
        
        # Aplicar la función a cada texto y crear la columna 'MENCIONES'
        df['MENCIONES'] = df['TEXTO_PLANO'].apply(encontrar_menciones_en_texto)
        
        # Si no hay menciones configuradas, asignar lista vacía a todas las filas
        if not lista_menciones:
            df['MENCIONES'] = [[] for _ in range(len(df))]
            logging.info("No hay menciones configuradas, asignando listas vacías")
        else:
            logging.info(f"Procesadas menciones para {len(df)} noticias")
        
        # Convertir listas a formato legible para Excel (opcional, solo para visualización)
        # Para la API esto no es necesario ya que se exporta como JSON
        # df['MENCIONES_EXCEL'] = df['MENCIONES'].apply(lambda x: ', '.join(x) if x else '')
        
        return df
        
    except Exception as e:
        logging.error(f"Error al buscar menciones en el DataFrame: {e}")
        # En caso de error, asignar lista vacía
        df['MENCIONES'] = [[] for _ in range(len(df))]
        return df

def detectar_crisis_por_tema(df_actual, df_historico=None, temas_fijos=None):
    """
    Detecta crisis basándose en 5+ noticias del mismo tema con valoración NEGATIVA.
    Una vez que un tema es crisis, todas las noticias de ese tema se marcan como crisis.
    
    Args:
        df_actual (DataFrame): DataFrame con las noticias actuales
        df_historico (DataFrame): DataFrame con noticias históricas (opcional)
        temas_fijos (list): Lista de temas fijos a excluir del análisis
    
    Returns:
        DataFrame: DataFrame con columna 'CRISIS' agregada
    """
    try:
        # Inicializar temas_fijos si no se proporciona
        if temas_fijos is None:
            temas_fijos = ["Actividades programadas"]
        
        # Combinar datos actuales e históricos
        df_completo = df_actual.copy()
        if df_historico is not None and not df_historico.empty:
            df_completo = pd.concat([df_historico, df_actual], ignore_index=True)
        
        # Filtrar solo noticias con valoración NEGATIVA (compatible con valores reales)
        df_negativas = df_completo[df_completo['VALORACION'] == 'NEGATIVO'].copy()
        
        # Excluir temas fijos
        df_negativas = df_negativas[~df_negativas['TEMA'].isin(temas_fijos)]
        
        # Contar noticias por tema
        conteo_temas = df_negativas['TEMA'].value_counts()
        
        # Identificar temas en crisis (5+ noticias negativas)
        temas_crisis = conteo_temas[conteo_temas >= 5].index.tolist()
        
        # Marcar crisis en el DataFrame actual
        df_actual['CRISIS'] = 'NO'
        
        # Marcar como crisis las noticias de temas en crisis
        for tema in temas_crisis:
            mask = df_actual['TEMA'] == tema
            df_actual.loc[mask, 'CRISIS'] = 'SI'
        
        logging.info(f"Detectados {len(temas_crisis)} temas en crisis: {temas_crisis}")
        logging.info(f"Marcadas {len(df_actual[df_actual['CRISIS'] == 'SI'])} noticias como crisis")
        
        return df_actual
        
    except Exception as e:
        logging.error(f"Error al detectar crisis por tema: {e}")
        # En caso de error, marcar todas como NO crisis
        df_actual['CRISIS'] = 'NO'
        return df_actual

def procesar_crisis_con_historico(df, historico_path, temas_fijos):
    """
    Procesa la detección de crisis cargando datos históricos y aplicando la lógica.
    
    Args:
        df (DataFrame): DataFrame con las noticias actuales
        historico_path (str): Ruta al archivo Excel con datos históricos
        temas_fijos (list): Lista de temas fijos a excluir
    
    Returns:
        DataFrame: DataFrame con columna 'CRISIS' procesada
    """
    try:
        # Cargar datos históricos si existe el archivo
        df_historico = None
        if os.path.exists(historico_path):
            df_historico = pd.read_excel(historico_path)
            logging.info(f"Datos históricos cargados: {len(df_historico)} noticias")
        else:
            logging.info("No se encontró archivo histórico. Procesando solo datos actuales.")
        
        # Aplicar detección de crisis
        df_procesado = detectar_crisis_por_tema(df, df_historico, temas_fijos)
        
        return df_procesado
        
    except Exception as e:
        logging.error(f"Error en procesamiento de crisis: {e}")
        df['CRISIS'] = 'NO'
        return df

def aplicar_heuristica_valoracion(valoracion_ia, texto, ministro_key_words, ministerios_key_words):
    """
    Aplica heurística de valoración basada en menciones del ministro/ministerios.
    
    Args:
        valoracion_ia (str): Valoración de la IA ("NEGATIVA", "NO_NEGATIVA", "OTRO")
        texto (str): Texto de la noticia
        ministro_key_words (str or list): Palabras clave para identificar al ministro
        ministerios_key_words (str or list): Palabras clave para identificar al ministerio
    
    Returns:
        str: Valoración final ("NEGATIVA", "POSITIVA", "NEUTRA")
    """
    # Si es negativa, mantener como negativa
    if valoracion_ia == "NEGATIVA":
        return "NEGATIVA"
    
    # Si NO es negativa, verificar menciones
    if valoracion_ia in ["NO_NEGATIVA", "OTRO"]:
        texto_lower = texto.lower()
        
        # Verificar si menciona a alguno de los ministros
        if ministro_key_words:
            if isinstance(ministro_key_words, list):
                # Si es una lista, buscar cualquiera de ellos (aplanar listas anidadas)
                for item in ministro_key_words:
                    if isinstance(item, list):
                        # Si es una lista anidada, procesar cada elemento
                        for ministro in item:
                            if ministro and ministro.lower() in texto_lower:
                                return "POSITIVA"
                    else:
                        # Si es un string directo
                        if item and item.lower() in texto_lower:
                            return "POSITIVA"
            else:
                # Si es un string, buscar directamente
                if ministro_key_words.lower() in texto_lower:
                    return "POSITIVA"
        
        # Verificar si menciona a alguno de los ministerios
        if ministerios_key_words:
            if isinstance(ministerios_key_words, list):
                # Si es una lista, buscar cualquiera de ellos (aplanar listas anidadas)
                for item in ministerios_key_words:
                    if isinstance(item, list):
                        # Si es una lista anidada, procesar cada elemento
                        for ministerio in item:
                            if ministerio and ministerio.lower() in texto_lower:
                                return "POSITIVA"
                    else:
                        # Si es un string directo
                        if item and item.lower() in texto_lower:
                            return "POSITIVA"
            else:
                # Si es un string, buscar directamente
                if ministerios_key_words.lower() in texto_lower:
                    return "POSITIVA"
        
        # Si no menciona a ninguno, es NEUTRA
        return "NEUTRA"
    
    # Para otros casos, mantener la valoración original
    return valoracion_ia

def get_gestion_from_html_obj(html_obj):
    """
    Extrae el campo GESTION basándose en indicadores técnicos en el HTML.
    Una nota es gestionada cuando aparece alguno de los indicadores técnicos.
    
    Args:
        html_obj: Objeto BeautifulSoup del HTML
        
    Returns:
        str: "GESTIONADA" si encuentra indicadores, "NO GESTIONADA" en caso contrario
    """
    if not html_obj:
        return "REBOTE"
    
    # Convertir todo el HTML a texto en minúsculas para búsqueda insensible a mayúsculas
    html_texto = html_obj.get_text().lower()
    
    # Lista de indicadores técnicos (en minúsculas para búsqueda)
    indicadores_gestion = [
        "content lab",
        "branded content", 
        "brand strategy",
        "brand studio",
        "special content",
        "contenido especial",
        "brandstudio",
        "brandstrategy",
        "contentlab",
        "brandedcontent",
        "contenidoespecial",
        "brand studio",
        "brand strategy"
    ]
    
    # Buscar cualquiera de los indicadores en el HTML
    for indicador in indicadores_gestion:
        if indicador in html_texto:
            logging.info(f"Indicador de gestión encontrado: '{indicador}'")
            return "GESTIONADA"
    
    # También buscar en atributos HTML (como class, id, etc.)
    for tag in html_obj.find_all():
        # Buscar en atributos del tag
        for attr_name, attr_value in tag.attrs.items():
            if isinstance(attr_value, str):
                attr_value_lower = attr_value.lower()
                for indicador in indicadores_gestion:
                    if indicador in attr_value_lower:
                        logging.info(f"Indicador de gestión encontrado en atributo {attr_name}: '{indicador}'")
                        return "GESTIONADA"
    
    return "REBOTE"

def marcar_o_valorar_con_ia(texto, funcion_ia, limite):
    """
    Función unificada para manejar el límite de texto en funciones de IA.
    Si el texto es muy largo o nulo, devuelve 'REVISAR MANUAL'.
    Si no, aplica la función IA y devuelve su resultado.
    
    Args:
        texto (str): Texto a procesar
        funcion_ia (function): Función de IA a aplicar (puede ser GPT u Ollama)
        limite (int): Límite de caracteres permitidos
        
    Returns:
        str: Resultado de la función IA o "REVISAR MANUAL"
    """
    if pd.isnull(texto) or len(texto) > limite:
        return "REVISAR MANUAL"
    return funcion_ia(texto)

def normalizar_medio(medio):
    """
    Normaliza el nombre del medio aplicando reglas automáticas.
    
    Args:
        medio (str): Nombre del medio a normalizar
        
    Returns:
        str: Nombre del medio normalizado
    """
    if not medio or pd.isnull(medio):
        return "Medio Desconocido"
    
    # 1. Convertir a string y limpiar espacios
    medio = str(medio).strip()
    
    # 2. Remover tildes y caracteres especiales
    medio = unicodedata.normalize('NFD', medio)
    medio = ''.join(c for c in medio if not unicodedata.combining(c))
    
    # 3. Remover extensiones de dominio
    if '.com' in medio.lower():
        medio = medio.split('.com')[0]
    if '.ar' in medio.lower():
        medio = medio.split('.ar')[0]
    if '.net' in medio.lower():
        medio = medio.split('.net')[0]
    
    # 4. Remover paréntesis y su contenido
    if '(' in medio:
        medio = medio.split('(')[0].strip()
    
    # 5. Separar camelCase automáticamente
    import re
    # Detectar camelCase y separar palabras
    medio = re.sub(r'([a-z])([A-Z])', r'\1 \2', medio)
    # Detectar números y separar
    medio = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', medio)
    
    # 6. Capitalizar primera letra de cada palabra
    medio = medio.title()
    
    # 7. Limpiar espacios múltiples
    medio = ' '.join(medio.split())
    
    return medio

def _normalizar_fecha_ddmmyyyy(fecha_raw: str) -> str | None:
    """
    Convierte 'd/m/yyyy', 'dd-mm-yyyy' o 'dd.mm.yyyy' a 'YYYY-MM-DD'.
    Acepta año de 2 o 4 dígitos; 2 dígitos se asumen 20xx si <=30, si no 19xx.
    """
    if not fecha_raw:
        return None
    try:
        import re
        txt = fecha_raw.strip().replace(" ", "")
        txt = re.sub(r"[\-.]", "/", txt)
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})", txt)
        if not m:
            # Intento ISO directo
            try:
                datetime.fromisoformat(txt)
                return txt
            except Exception:
                return None
        d, mth, y = m.groups()
        if len(y) == 2:
            yi = int(y)
            y = f"20{y}" if yi <= 30 else f"19{y}"
        return f"{y}-{mth.zfill(2)}-{d.zfill(2)}"
    except Exception:
        return None

def cargar_temas_desde_txt(path_txt: str) -> tuple[list, dict]:
    """
    Carga temas desde un TXT (uno por línea). Soporta líneas con fecha opcional: 'Tema; dd/mm/yyyy'.

    Returns:
        (lista_temas, tema_a_fecha_iso)
    """
    temas: list[str] = []
    tema_a_fecha: dict[str, str] = {}
    try:
        import os
        import re
        if not os.path.exists(path_txt):
            logging.warning(f"No se encontró {path_txt}. Usando lista mínima.")
            return ["Actividades programadas"], {}

        pat = re.compile(r"^(?P<tema>.+?)[\s;:|\t\-—–]*?(?P<fecha>\d{1,2}[\/.\-]\d{1,2}[\/.\-]\d{2,4})?\s*$")

        with open(path_txt, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line:
                    continue
                m = pat.match(line)
                if not m:
                    continue
                nombre = m.group("tema").strip()
                fecha = m.group("fecha")

                if nombre and nombre not in temas:
                    temas.append(nombre)
                if fecha:
                    iso = _normalizar_fecha_ddmmyyyy(fecha)
                    if iso:
                        tema_a_fecha[nombre] = iso

        logging.info(f"Temas cargados desde TXT: {len(temas)} | con fecha: {len(tema_a_fecha)}")
        return temas, tema_a_fecha
    except Exception as e:
        logging.error(f"Error cargando temas desde TXT: {e}")
        return ["Actividades programadas"], {}

def validar_urls_ejes(urls: list) -> dict:
    """
    Valida que las URLs pertenezcan al dominio ejes.com
    
    Args:
        urls (list): Lista de URLs a validar
        
    Returns:
        dict: Diccionario con URLs válidas, no válidas y estadísticas
        {
            "validas": [urls_validas],
            "no_validas": [urls_no_validas],
            "motivos": {url: motivo},
            "estadisticas": {
                "total": int,
                "validas": int,
                "no_validas": int
            }
        }
    """
    try:
        urls_validas = []
        urls_no_validas = []
        motivos = {}
        
        for url in urls:
            # Verificar si es None o string vacío
            if url is None:
                urls_no_validas.append(str(url))
                motivos[str(url)] = "URL es None"
                continue
                
            # Convertir a string si no lo es
            if not isinstance(url, str):
                urls_no_validas.append(str(url))
                motivos[str(url)] = f"URL no es string (tipo: {type(url).__name__})"
                continue
            
            # Verificar si es string vacío
            if url.strip() == "":
                urls_no_validas.append(url)
                motivos[url] = "URL vacía"
                continue
            
            # Verificar que contenga ejes.com (case-insensitive)
            if 'ejes.com' not in url.lower():
                urls_no_validas.append(url)
                motivos[url] = "URL no pertenece a ejes.com"
                continue
            
            # Verificar formato básico de URL
            if not url.startswith(('http://', 'https://')):
                urls_no_validas.append(url)
                motivos[url] = "URL no tiene protocolo http/https"
                continue
            
            # Si pasa todas las validaciones, es válida
            urls_validas.append(url)
        
        estadisticas = {
            "total": len(urls),
            "validas": len(urls_validas),
            "no_validas": len(urls_no_validas)
        }
        
        logging.info(f"Validación de URLs: {estadisticas['validas']}/{estadisticas['total']} válidas")
        
        if urls_no_validas:
            for url in urls_no_validas:
                logging.warning(f"URL no válida: {url} - Motivo: {motivos[url]}")
        
        return {
            "validas": urls_validas,
            "no_validas": urls_no_validas,
            "motivos": motivos,
            "estadisticas": estadisticas
        }
        
    except Exception as e:
        logging.error(f"Error en validación de URLs: {e}")
        # En caso de error, convertir todas las URLs a string para evitar problemas de hash
        urls_str = [str(url) if url is not None else "None" for url in urls]
        return {
            "validas": [],
            "no_validas": urls_str,
            "motivos": {url: f"Error en validación: {str(e)}" for url in urls_str},
            "estadisticas": {
                "total": len(urls),
                "validas": 0,
                "no_validas": len(urls)
            }
        }
 
 
